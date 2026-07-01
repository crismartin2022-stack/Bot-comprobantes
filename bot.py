import os
import json
import base64
import logging
import asyncio
from datetime import datetime, timezone, timedelta
from io import BytesIO

import httpx
import anthropic
import redis.asyncio as aioredis
from apscheduler.schedulers.asyncio import AsyncIOScheduler
from apscheduler.triggers.cron import CronTrigger
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup, ReactionTypeEmoji
from telegram.error import RetryAfter, TimedOut, NetworkError
from telegram.ext import (
    Application, CommandHandler, MessageHandler,
    filters, ContextTypes, CallbackQueryHandler
)

# ── Logging ───────────────────────────────────────────────────────────────────
logging.basicConfig(format="%(asctime)s | %(levelname)s | %(message)s", level=logging.INFO)
log = logging.getLogger(__name__)
logging.getLogger("httpx").setLevel(logging.WARNING)
logging.getLogger("httpcore").setLevel(logging.WARNING)


# ── Helper flood control ──────────────────────────────────────────────────────
async def send_safe(coro_fn, retries=3):
    for attempt in range(retries):
        try:
            return await coro_fn()
        except RetryAfter as e:
            wait = e.retry_after + 1
            log.warning(f"Flood control: esperando {wait}s (intento {attempt+1})")
            await asyncio.sleep(wait)
        except (TimedOut, NetworkError) as e:
            log.warning(f"Error de red Telegram: {e} (intento {attempt+1})")
            await asyncio.sleep(2)
        except Exception as e:
            log.error(f"Error enviando mensaje Telegram: {e}")
            return None
    return None

async def reaccionar(bot, chat_id: int, msg_id: int, emoji: str):
    async with _semaforo_reacciones:
        for attempt in range(3):
            try:
                await bot.set_message_reaction(
                    chat_id=chat_id,
                    message_id=msg_id,
                    reaction=[ReactionTypeEmoji(emoji=emoji)]
                )
                await asyncio.sleep(2)
                return
            except RetryAfter as e:
                wait = e.retry_after + 1
                log.warning(f"Reacción {emoji} flood control: esperando {wait}s")
                await asyncio.sleep(wait)
            except Exception as e:
                log.warning(f"Reacción {emoji} no aplicada: {e}")
                return

# ── Config ────────────────────────────────────────────────────────────────────
TELEGRAM_TOKEN = os.environ["TELEGRAM_TOKEN"]
ANTHROPIC_KEY  = os.environ["ANTHROPIC_API_KEY"]
ADMIN_ID       = 531707598
ARG_TZ         = timezone(timedelta(hours=-3))
claude         = anthropic.Anthropic(api_key=ANTHROPIC_KEY)
REDIS_URL      = os.environ.get("REDIS_URL", "")
redis_client   = None  # Se inicializa en main_async

# ── Storage ───────────────────────────────────────────────────────────────────
store: dict = {}
pendientes: dict = {}
esperando_pie: dict = {}
mensajes_rechazo: dict = {}
received_log: dict = {}  # {chat_id: [{"msg_id", "fecha", "remitente", "estado"}]}
semaforo_claude = asyncio.Semaphore(10)  # Máximo 3 análisis simultáneos
cola_procesamiento = asyncio.Queue()  # Cola local (fallback si no hay Redis)
_github_backup_counter = 0  # Contador para throttle de backup GitHub
_semaforo_reacciones = asyncio.Semaphore(1)  # Máximo 1 reacción a la vez

DATA_FILE    = "/data/store.json"       # Volume de Railway (persistente)
LOG_FILE     = "/data/received_log.json"  # Log de imágenes recibidas
QUEUE_FILE   = "/data/cola_pendiente.json"  # Cola persistente
GITHUB_TOKEN = os.environ.get("GITHUB_TOKEN", "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")
CLOUDINARY_CLOUD_NAME = os.environ.get("CLOUDINARY_CLOUD_NAME", "")
CLOUDINARY_API_KEY    = os.environ.get("CLOUDINARY_API_KEY", "")
CLOUDINARY_API_SECRET = os.environ.get("CLOUDINARY_API_SECRET", "")
GITHUB_REPO  = "crismartin2022-stack/Bot-comprobantes"
GITHUB_FILE  = "store.json"
GITHUB_API   = f"https://api.github.com/repos/{GITHUB_REPO}/contents/{GITHUB_FILE}"
GITHUB_HEADERS = {
    "Authorization": f"token {GITHUB_TOKEN}",
    "Accept": "application/vnd.github.v3+json",
}

def guardar_store():
    """Guarda el store en disco (Railway Volume) y en GitHub como respaldo."""
    try:
        os.makedirs(os.path.dirname(DATA_FILE), exist_ok=True)
        store_limpio = {}
        for cid, datos in store.items():
            store_limpio[cid] = {k: v for k, v in datos.items()
                                  if k not in ("_task",) and not isinstance(v, bytes)}
        with open(DATA_FILE, "w", encoding="utf-8") as f:
            json.dump(store_limpio, f, ensure_ascii=False, indent=2)
        log.info("Store guardado en Volume ✅")
        # Guardar cola pendiente
        try:
            guardar_cola()
        except Exception as qe:
            log.error(f"Error guardando cola: {qe}")
        # Guardar log de imágenes
        try:
            with open(LOG_FILE, "w", encoding="utf-8") as f:
                json.dump(received_log, f, ensure_ascii=False, indent=2)
        except Exception as le:
            log.error(f"Error guardando log: {le}")
        # Respaldo en GitHub cada 10 guardados
        global _github_backup_counter
        _github_backup_counter += 1
        if _github_backup_counter >= 10:
            _github_backup_counter = 0
            import threading
            threading.Thread(target=_guardar_github_sync, args=(store_limpio,), daemon=True).start()
    except Exception as e:
        log.error(f"Error guardando store: {e}")

def _guardar_github_sync(store_limpio: dict):
    """Respaldo asíncrono en GitHub."""
    try:
        contenido = json.dumps(store_limpio, ensure_ascii=False, indent=2)
        b64 = base64.b64encode(contenido.encode("utf-8")).decode("utf-8")
        r = httpx.get(GITHUB_API, headers=GITHUB_HEADERS, timeout=10)
        sha = r.json().get("sha") if r.status_code == 200 else None
        payload = {"message": f"backup {datetime.now().strftime('%d/%m/%Y %H:%M')}", "content": b64}
        if sha:
            payload["sha"] = sha
        resp = httpx.put(GITHUB_API, headers=GITHUB_HEADERS, json=payload, timeout=15)
        if resp.status_code in (200, 201):
            log.info("Respaldo GitHub ✅")
        else:
            log.error(f"Error respaldo GitHub: {resp.status_code}")
    except Exception as e:
        log.error(f"Error respaldo GitHub: {e}")


async def subir_cloudinary(image_bytes: bytes, mime: str, public_id: str = None) -> str:
    if not CLOUDINARY_CLOUD_NAME or not CLOUDINARY_API_KEY or not CLOUDINARY_API_SECRET:
        return ""
    try:
        import hashlib, time
        timestamp = str(int(time.time()))
        params = f"folder=comprobantes&timestamp={timestamp}"
        if public_id:
            params = f"folder=comprobantes&public_id={public_id}&timestamp={timestamp}"
        signature = hashlib.sha1(f"{params}{CLOUDINARY_API_SECRET}".encode()).hexdigest()
        ext = "jpg" if "jpeg" in mime else mime.split("/")[-1]
        data = {"api_key": CLOUDINARY_API_KEY, "timestamp": timestamp, "signature": signature, "folder": "comprobantes"}
        if public_id:
            data["public_id"] = public_id
        files = {"file": (f"comprobante.{ext}", image_bytes, mime)}
        async with httpx.AsyncClient(timeout=20) as client:
            resp = await client.post(f"https://api.cloudinary.com/v1_1/{CLOUDINARY_CLOUD_NAME}/image/upload", files=files, data=data)
        if resp.status_code == 200:
            url = resp.json().get("secure_url", "")
            log.info(f"Cloudinary ✅ {url}")
            return url
        else:
            log.error(f"Cloudinary error {resp.status_code}: {resp.text[:300]}")
            return ""
    except Exception as e:
        log.error(f"Error subiendo a Cloudinary: {e}")
        return ""

def cargar_store():
    """Carga el store desde Volume. Si no existe, intenta desde GitHub."""
    global store, received_log
    # Cargar log de imágenes
    try:
        if os.path.exists(LOG_FILE):
            with open(LOG_FILE, "r", encoding="utf-8") as f:
                received_log = json.load(f)
            log.info(f"Log cargado: {sum(len(v) for v in received_log.values())} entradas")
    except Exception as e:
        log.error(f"Error cargando log: {e}")
        received_log = {}
    try:
        if os.path.exists(DATA_FILE):
            with open(DATA_FILE, "r", encoding="utf-8") as f:
                store = json.load(f)
            total = sum(len(d.get("registros", [])) for d in store.values())
            log.info(f"Store cargado desde Volume: {len(store)} grupos, {total} registros ✅")
            return
    except Exception as e:
        log.error(f"Error cargando desde Volume: {e}")

    # Fallback: cargar desde GitHub
    try:
        r = httpx.get(GITHUB_API, headers=GITHUB_HEADERS, timeout=10)
        if r.status_code == 200:
            contenido_b64 = r.json().get("content", "")
            contenido = base64.b64decode(contenido_b64).decode("utf-8")
            store = json.loads(contenido)
            total = sum(len(d.get("registros", [])) for d in store.values())
            log.info(f"Store cargado desde GitHub: {len(store)} grupos, {total} registros ✅")
        else:
            log.info("Sin store previo, iniciando vacío")
            store = {}
    except Exception as e:
        log.error(f"Error cargando desde GitHub: {e}")
        store = {}

def guardar_cola():
    """Guarda los items pendientes de la cola en disco."""
    try:
        items = []
        # Vaciar y rellenar para inspeccionar sin consumir
        temp = []
        while not cola_procesamiento.empty():
            try:
                item = cola_procesamiento.get_nowait()
                # No guardar image_bytes (muy pesado) — solo metadata
                items.append({
                    "chat_id": item["chat_id"],
                    "msg_id": item["msg_id"],
                    "pie": item.get("pie", ""),
                    "nombre_g": item["nombre_g"],
                    "origen": item.get("origen", "grupo"),
                    "fecha": now_arg().strftime("%d/%m/%Y %H:%M"),
                    "_sin_imagen": True  # Marcador: imagen no guardada
                })
                temp.append(item)
                cola_procesamiento.task_done()
            except Exception:
                break
        # Restaurar items
        for item in temp:
            cola_procesamiento.put_nowait(item)
        with open(QUEUE_FILE, "w", encoding="utf-8") as f:
            json.dump(items, f, ensure_ascii=False, indent=2)
        log.info(f"Cola guardada: {len(items)} items pendientes")
    except Exception as e:
        log.error(f"Error guardando cola: {e}")

def get_store(chat_id: int, nom: str = "") -> dict:
    cid = str(chat_id)
    if cid not in store:
        store[cid] = {
            "nombre": nom or f"Grupo {cid}",
            "semana_actual": semana_label(),
            "registros": [],
            "errores": [],
            "chat_id": chat_id,
            "ultimo_resumen_idx": 0,
            "total_mensual": 0.0,
            "mes_actual": datetime.now(ARG_TZ).strftime("%m/%Y"),
        }
        guardar_store()
    if nom and store[cid]["nombre"] != nom:
        store[cid]["nombre"] = nom
    # Resetear total mensual si cambió el mes
    mes_ahora = datetime.now(ARG_TZ).strftime("%m/%Y")
    if store[cid].get("mes_actual") != mes_ahora:
        store[cid]["total_mensual"] = 0.0
        store[cid]["mes_actual"]    = mes_ahora
    return store[cid]

def semana_label() -> str:
    return f"Semana {datetime.now(ARG_TZ).strftime('%d/%m/%Y')}"

def now_arg() -> datetime:
    return datetime.now(ARG_TZ)

def get_nombre_grupo(update: Update) -> str:
    chat = update.effective_chat
    return "Privado" if chat.type == "private" else (chat.title or f"Grupo {chat.id}")

def grupos_disponibles() -> list:
    return [(cid, d["nombre"]) for cid, d in store.items() if d.get("chat_id") != ADMIN_ID]

# ── Análisis con Claude ───────────────────────────────────────────────────────
SYSTEM_PROMPT = """Eres un asistente experto en análisis de comprobantes bancarios argentinos.
Analizá la imagen y respondé ÚNICAMENTE con un JSON válido sin backticks ni markdown.
{
  "fecha": "DD/MM/YYYY",
  "hora": "HH:MM o vacío",
  "tipo": "TRF / DEP / PAGO / otro",
  "monto": número sin símbolos (ej: 15000.00),
  "moneda": "ARS / USD / otro",
  "remitente": "nombre completo del que envía según el comprobante, o vacío",
  "remitente_cuil": "CUIL/DNI del REMITENTE/ORIGEN. Buscarlo en: campo CUIL, DNI, CUIT del que ENVÍA el dinero. Puede aparecer como 'CUIL', 'CUIT/CUIL', 'DNI', 'Documento' bajo la sección 'De', 'Desde', 'Origen', 'Cuenta origen', 'Enviaste'. Formato con guiones (ej: 20-12345678-9). NUNCA tomar el CUIL del destinatario. Si no se encuentra dejar VACÍO.",
  "destinatario": "nombre completo del que recibe o vacío",
  "banco_origen": "banco origen o vacío",
  "banco_destino": "banco destino o vacío",
  "referencia": "identificador único de la transacción. Buscar en TODOS estos campos y tomar el primero que aparezca: 'Número de operación', 'N° operación', 'Id Op', 'ID operación', 'Coelsa ID', 'Código de identificación', 'ID transacción', 'Referencia', 'Comprobante N°', 'N° comprobante', 'Transaction ID', 'Nro. de transacción', 'Número de comprobante', 'Cod. operación', 'Número de seguimiento'. Es el código alfanumérico único que identifica esta transacción específica. SIEMPRE extraerlo si está visible.",
  "concepto": "concepto o vacío",
  "estado": "Exitoso / Pendiente / Rechazado",
  "cvu_ultimos4": "últimos 4 dígitos del CVU/CBU del RECEPTOR/DESTINATARIO. Buscarlo en: campo CVU, CBU, Cuenta Receptor, Cuenta Destino, Cuenta, número de cuenta del que RECIBE el dinero. Es un número largo de 22 dígitos — tomar los ÚLTIMOS 4. Si no se encuentra dejar VACÍO (nunca inventar)",
  "tiene_remitente": true o false según si el comprobante tiene datos del remitente visibles,
  "notas": "cualquier dato relevante adicional"
}
IMPORTANTE sobre el CUIL del remitente:
- Mercado Pago: está bajo la sección 'De' junto al nombre del que envía
- Naranja X: campo 'CUIL' bajo 'Cuenta origen'
- Billetera País / Banco: campo 'CUIL' o 'CUIT/CUIL' de la cuenta origen
- Si el comprobante muestra DOS CUIL (origen y destino), tomar SIEMPRE el del ORIGEN/REMITENTE

IMPORTANTE sobre el CVU/CBU receptor:
- Mercado Pago: campo 'CVU' bajo 'Para'
- Billetera País: campo 'Cuenta Receptor'  
- Banco tradicional: campo 'CBU destino' o 'Cuenta destino'
- Naranja X, Ualá, etc: campo 'CVU' o 'Cuenta'
Siempre buscar el número largo (22 dígitos) asociado al RECEPTOR y tomar los últimos 4."""

async def analizar_imagen(image_bytes: bytes, mime: str, reintentos: int = 3) -> dict:
    b64 = base64.standard_b64encode(image_bytes).decode("utf-8")
    ultimo_error = None
    for intento in range(1, reintentos + 1):
        try:
            resp = await asyncio.wait_for(
                asyncio.get_event_loop().run_in_executor(
                    None,
                    lambda: claude.messages.create(
                        model="claude-sonnet-4-6",
                        max_tokens=1000,
                        system=[{"type": "text", "text": SYSTEM_PROMPT, "cache_control": {"type": "ephemeral"}}],
                        messages=[{"role": "user", "content": [
                            {"type": "image", "source": {"type": "base64", "media_type": mime, "data": b64}},
                            {"type": "text", "text": "Analizá este comprobante bancario argentino."}
                        ]}]
                    )
                ),
                timeout=30
            )
            text = resp.content[0].text
            try:
                return json.loads(text.replace("```json", "").replace("```", "").strip())
            except Exception:
                return {"notas": text, "estado": "Error al parsear", "cvu_ultimos4": "", "tiene_remitente": False}
        except asyncio.TimeoutError:
            ultimo_error = "Timeout de 30 segundos"
            log.warning(f"Intento {intento}/{reintentos} timeout")
            if intento < reintentos:
                await asyncio.sleep(2)
        except Exception as e:
            ultimo_error = e
            log.warning(f"Intento {intento}/{reintentos} fallido: {e}")
            if intento < reintentos:
                await asyncio.sleep(2)
    log.error(f"Todos los reintentos fallaron: {ultimo_error}")
    return {"notas": str(ultimo_error), "estado": "Error timeout", "cvu_ultimos4": "", "tiene_remitente": False}

def normalizar(texto: str) -> str:
    """Normaliza texto para comparación: minúsculas, sin espacios extra, sin puntos."""
    return " ".join(texto.lower().replace(".", "").replace("-", "").split())

def similitud(a: str, b: str) -> float:
    """Calcula similitud entre dos strings (0 a 1). Usa distancia de edición simple."""
    if not a or not b:
        return 0.0
    a, b = a.lower(), b.lower()
    if a == b:
        return 1.0
    # Distancia de Levenshtein simplificada
    la, lb = len(a), len(b)
    if abs(la - lb) > max(la, lb) * 0.5:
        return 0.0
    dp = list(range(lb + 1))
    for i, ca in enumerate(a):
        nuevo = [i + 1]
        for j, cb in enumerate(b):
            nuevo.append(min(dp[j] + (0 if ca == cb else 1), dp[j+1] + 1, nuevo[j] + 1))
        dp = nuevo
    distancia = dp[lb]
    return 1 - distancia / max(la, lb)

def palabra_en_texto(palabra: str, texto: str, umbral: float = 0.82) -> bool:
    """Verifica si una palabra está en el texto con tolerancia a errores de tipeo."""
    if palabra in texto:
        return True
    # Buscar palabra similar en cada palabra del texto
    for w in texto.split():
        if similitud(palabra, w) >= umbral:
            return True
    return False

def extraer_partes(nombre: str) -> tuple[list, str]:
    """Devuelve (lista_de_nombres, apellido) de un nombre completo normalizado."""
    partes = normalizar(nombre).split()
    if not partes:
        return [], ""
    apellido = partes[-1]
    nombres  = partes[:-1]
    return nombres, apellido

def verificar_pie(resultado: dict, pie: str) -> tuple[bool, str]:
    """
    El pie siempre es la fuente correcta del titular.
    Solo rechaza si el ID de operación está repetido (manejado en es_duplicado).
    Si hay pie, acepta y usa esos datos. Si no hay pie, acepta igual.
    Hace una comparación informativa entre imagen y pie pero NO rechaza por diferencia de nombres.
    """
    if not pie:
        return True, ""

    # Si no hay datos del remitente en la imagen, usar el pie directamente
    if not resultado.get("tiene_remitente") or not (resultado.get("remitente") or "").strip():
        return True, "sin_datos_imagen"

    # Hay datos en imagen y pie — comparar informativamente pero siempre aceptar
    return True, "coincide"

# ── Generador de Excel ────────────────────────────────────────────────────────
def generar_excel(registros: list, semana: str, nombre: str, es_errores: bool = False, duplicados: list = None) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Rechazados" if es_errores else "Comprobantes"

    header_fill  = PatternFill("solid", fgColor="8B0000" if es_errores else "4B0082")
    error_fill   = PatternFill("solid", fgColor="FF4444")
    ok_fill      = PatternFill("solid", fgColor="1A5C2A")
    header_font  = Font(bold=True, color="FFFFFF", size=11)
    border = Border(left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="thin"),  bottom=Side(style="thin"))

    if es_errores:
        headers    = ["#", "GRUPO", "FECHA", "TIPO", "TITULAR CTA", "REMITENTE COMPROBANTE",
                      "REMITENTE PIE", "MONTO", "CUENTA (CVU)", "MOTIVO", "Fecha Carga"]
        col_widths = [4, 20, 13, 10, 28, 28, 28, 14, 14, 40, 16]
    else:
        headers    = ["#", "GRUPO", "FECHA DE ENVIO", "TRF O DEPOSITO", "TITULAR DE LA CTA",
                      "FECHA TICKET", "HORA TICKET", "CUENTA (CVU)", "MONTO",
                      "Remitente", "CUIL Remitente", "Banco Origen", "Estado", "Origen", "Notas", "Imagen"]
        col_widths = [4, 20, 22, 12, 30, 13, 11, 14, 14, 25, 20, 18, 11, 10, 28, 45]

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 20

    sin_cvu = 0
    for i, r in enumerate(registros, 2):
        cvu = (r.get("cvu_ultimos4") or "").strip()
        tiene_cvu = bool(cvu)
        if not tiene_cvu and not es_errores:
            sin_cvu += 1

        if es_errores:
            fila = [
                i - 1, nombre,
                r.get("fecha", ""),
                r.get("tipo", "TRF"),
                r.get("destinatario", ""),
                r.get("remitente", ""),
                r.get("_pie", ""),
                r.get("monto", ""),
                cvu if tiene_cvu else "⚠️ SIN CVU",
                r.get("_motivo_error", ""),
                r.get("_fecha_carga", ""),
            ]
        else:
            fila = [
                i - 1, nombre, semana,
                r.get("tipo", "TRF"),
                r.get("destinatario", ""),
                r.get("fecha", ""),
                r.get("hora", ""),
                cvu if tiene_cvu else "⚠️ SIN CVU",
                r.get("monto", ""),
                r.get("remitente", ""),
                r.get("remitente_cuil", ""),
                r.get("banco_origen", ""),
                r.get("estado", ""),
                r.get("_origen", "grupo"),
                r.get("notas", ""),
                r.get("_imagen_url", ""),
            ]

        cvu_col = 9 if es_errores else 8
        img_col = len(fila) if not es_errores else None
        for col, val in enumerate(fila, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=(col == len(fila)))
            if col == cvu_col:
                cell.fill = ok_fill if tiene_cvu else error_fill
                cell.font = Font(bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            if not es_errores and img_col and col == img_col and val:
                cell.hyperlink = val
                cell.value = "Ver imagen"
                cell.font = Font(color="0563C1", underline="single")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    # Hoja de duplicados (en archivo de rechazados)
    if es_errores and duplicados:
        ws_dup = wb.create_sheet("Duplicados")
        dup_headers    = ["#", "FECHA", "HORA", "REMITENTE", "MONTO", "REFERENCIA", "Fecha Carga"]
        dup_col_widths = [4, 13, 10, 30, 14, 35, 16]
        dup_fill = PatternFill("solid", fgColor="FF6600")
        for col, (h, w) in enumerate(zip(dup_headers, dup_col_widths), 1):
            cell = ws_dup.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = dup_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
            ws_dup.column_dimensions[get_column_letter(col)].width = w
        for i, r in enumerate(duplicados, 2):
            fila_dup = [
                i - 1,
                r.get("fecha", ""),
                r.get("hora", ""),
                r.get("remitente", ""),
                r.get("monto", ""),
                r.get("referencia", ""),
                r.get("_fecha_carga", ""),
            ]
            for col, val in enumerate(fila_dup, 1):
                cell = ws_dup.cell(row=i, column=col, value=val)
                cell.border = border
                cell.alignment = Alignment(vertical="center")
        ws_dup.freeze_panes = "A2"

    # Hoja resumen
    ws2 = wb.create_sheet("Resumen")
    for col, h in enumerate(["Grupo", "Semana", "Comprobantes", "Con CVU", "Sin CVU", "Total ARS"], 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
    total = sum(float(r.get("monto") or 0) for r in registros)
    ws2.append([nombre, semana, len(registros), len(registros) - sin_cvu, sin_cvu, total])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

# ── Formateo ──────────────────────────────────────────────────────────────────
def formatear_resultado(r: dict, num: int, grupo: str = "", pie: str = "") -> str:
    cvu = (r.get("cvu_ultimos4") or "").strip()
    cvu_line = f"🏦 CVU (últimos 4): `****{cvu}`" if cvu else "🔴 *CVU NO ENCONTRADO*"
    monto = r.get("monto")
    monto_fmt = f"${float(monto):,.0f}" if monto else "—"
    grupo_line = f"📍 Grupo: *{grupo}*\n" if grupo else ""
    pie_line   = f"📝 Pie: _{pie}_\n" if pie else ""
    return (
        f"✅ *Comprobante #{num} procesado*\n"
        f"─────────────────────\n"
        f"{grupo_line}{pie_line}"
        f"📅 {r.get('fecha','—')} ⏰ {r.get('hora','—')}\n"
        f"💸 {r.get('tipo','—')} | 💰 {monto_fmt} {r.get('moneda','ARS')}\n"
        f"👤 Remitente: {r.get('remitente','—')}\n"
        f"👤 Destinatario: {r.get('destinatario','—')}\n"
        f"🏛 Banco: {r.get('banco_origen','—')}\n"
        f"{cvu_line}\n"
        f"📋 Estado: {r.get('estado','—')}\n"
    )

# ── Procesar comprobante con verificación de pie ──────────────────────────────
async def procesar_comprobante(image_bytes: bytes, mime: str, pie: str,
                                chat_id: int, nombre_g: str, origen: str,
                                bot, chat_msg_id: int):
    """Analiza imagen, verifica pie y guarda en registro correcto o errores."""
    datos = get_store(chat_id, nombre_g)
    # Registrar imagen recibida
    from telegram import Update as _Update
    entrada_log = {
        "msg_id": chat_msg_id,
        "fecha": now_arg().strftime("%d/%m/%Y %H:%M"),
        "nombre_g": nombre_g,
        "estado": "procesando",
        "pie": pie or "",
    }
    cid_str = str(chat_id)
    if cid_str not in received_log:
        received_log[cid_str] = []
    received_log[cid_str].append(entrada_log)

    async with semaforo_claude:
        resultado = await analizar_imagen(image_bytes, mime)

    # Si el comprobante no tiene datos del remitente, usar el pie
    if pie and not resultado.get("tiene_remitente"):
        resultado["remitente"] = pie
        resultado["_pie_como_fuente"] = True

    # Extraer CUIL/DNI del pie si no se encontró en la imagen
    if pie and not (resultado.get("remitente_cuil") or "").strip():
        import re
        # Buscar patrones: 20-12345678-9 o 20123456789 o DNI 12345678
        cuil_match = re.search(r'\b(\d{2}[-.]?\d{7,8}[-.]?\d{1})\b', pie)
        dni_match  = re.search(r'\b(\d{7,8})\b', pie)
        if cuil_match:
            resultado["remitente_cuil"] = cuil_match.group(1)
        elif dni_match:
            resultado["remitente_cuil"] = dni_match.group(1)

    coincide, motivo = verificar_pie(resultado, pie)
    num = len(datos["registros"]) + len(datos["errores"]) + 1
    resultado["_num"] = num
    resultado["_fecha_carga"] = now_arg().strftime("%d/%m/%Y %H:%M")
    resultado["_origen"] = origen
    resultado["_pie"] = pie or ""
    resultado["_msg_id"] = chat_msg_id

    # Verificar si hubo error de timeout en el análisis
    if resultado.get("estado") == "Error timeout":
        resultado["_motivo_error"] = "Error de conexión con IA — reintentá más tarde"
        datos["errores"].append(resultado)
        guardar_store()
        await send_safe(lambda: bot.send_message(
            chat_id=chat_id,
            text=f"⚠️ Comprobante #{num} no procesado — error de conexión. Reenvialo.",
            reply_to_message_id=chat_msg_id
        ))
        asyncio.create_task(reaccionar(bot, chat_id, chat_msg_id, "🤡"))
        return

    if coincide:
        # Verificar duplicado
        if es_duplicado(resultado, datos["registros"]):
            resultado["_duplicado"] = True
            datos.setdefault("duplicados", []).append(resultado)
            guardar_store()
            monto = resultado.get("monto")
            monto_fmt = f"${float(monto):,.0f}" if monto else "—"
            ref_nuevo = (resultado.get("referencia") or "").strip()
            msg_id_original = None
            for r in datos["registros"]:
                if (r.get("referencia") or "").strip() == ref_nuevo:
                    msg_id_original = r.get("_msg_id")
                    break
            texto_dup = f"🤨 *Duplicado* — msg #{chat_msg_id}"
            if msg_id_original:
                texto_dup += f" ya procesado en msg #{msg_id_original}"
            texto_dup += f"\n💰 {monto_fmt} | 👤 {resultado.get('remitente','—')}"
            _cid, _mid, _txt = chat_id, chat_msg_id, texto_dup
            await send_safe(lambda: bot.send_message(chat_id=_cid, text=_txt, parse_mode="Markdown", reply_to_message_id=_mid))
            asyncio.create_task(reaccionar(bot, _cid, _mid, "🤨"))
            return
        datos["registros"].append(resultado)
        guardar_store()
        asyncio.create_task(_subir_imagen_cloudinary(image_bytes, mime, resultado, datos))
        for entry in received_log.get(str(chat_id), []):
            if entry.get("msg_id") == chat_msg_id:
                entry["estado"] = "procesada"
                break
        cvu = (resultado.get("cvu_ultimos4") or "").strip()
        monto = resultado.get("monto")
        monto_fmt = f"${float(monto):,.0f}" if monto else "—"
        remitente = resultado.get("remitente") or "—"
        sin_datos = not resultado.get("tiene_remitente") and not (resultado.get("remitente") or "").strip()
        if not cvu or sin_datos:
            _cid, _mid, _num = chat_id, chat_msg_id, num
            _motivos = (["sin CVU/CBU"] if not cvu else []) + (["sin datos de remitente"] if sin_datos else [])
            _motivos_txt = ", ".join(_motivos)
            asyncio.create_task(reaccionar(bot, _cid, _mid, "🤔"))
            await send_safe(lambda: bot.send_message(chat_id=_cid, text=f"🤔 Comprobante #{_num} aprobado pero con observaciones: {_motivos_txt}\n💰 {monto_fmt}", reply_to_message_id=_mid))
            await send_safe(lambda: bot.send_message(chat_id=ADMIN_ID, text=f"⚠️ *Comprobante con observaciones — {nombre_g}*\n#{_num} | {_motivos_txt}\n👤 {remitente}\n💰 {monto_fmt}\n📅 {resultado.get('fecha','—')}", parse_mode="Markdown"))
        else:
            asyncio.create_task(reaccionar(bot, chat_id, chat_msg_id, "👍"))
    else:
        resultado["_motivo_error"] = motivo
        datos["errores"].append(resultado)
        guardar_store()
        monto = resultado.get("monto")
        monto_fmt = f"${float(monto):,.0f}" if monto else "—"
        texto_error = (
            f"⛔ *Comprobante #{num} RECHAZADO — {nombre_g}*\n"
            f"─────────────────────\n"
            f"💰 Monto: {monto_fmt}\n"
            f"❌ {motivo}\n\n"
            f"_Avisá en el grupo para que corrijan los datos._"
        )
        _cid, _mid, _num, _txt_err = chat_id, chat_msg_id, num, texto_error
        await send_safe(lambda: bot.send_message(chat_id=ADMIN_ID, text=_txt_err, parse_mode="Markdown"))
        sent = await send_safe(lambda: bot.send_message(chat_id=_cid, text=f"⛔ Comprobante #{_num} rechazado — datos no coinciden.\nPor favor corregí respondiendo a este mensaje con el nombre y CUIL correcto.", reply_to_message_id=_mid))
        if sent:
            mensajes_rechazo[(_cid, sent.message_id)] = {"num": _num, "chat_id": _cid}
        asyncio.create_task(reaccionar(bot, _cid, _mid, "❌"))


async def _subir_imagen_cloudinary(image_bytes: bytes, mime: str, resultado: dict, datos: dict):
    num = resultado.get("_num", "x")
    nombre = datos.get("nombre", "grupo").replace(" ", "_").replace("-", "_")
    fecha = resultado.get("_fecha_carga", "").replace("/", "-").replace(" ", "_").replace(":", "-")
    public_id = f"{nombre}_num{num}_{fecha}"
    url = await subir_cloudinary(image_bytes, mime, public_id)
    if url:
        resultado["_imagen_url"] = url
        guardar_store()

# ── Tarea para procesar foto sin pie después de 5 segundos ──────────────────
async def procesar_sin_pie(key, app):
    await asyncio.sleep(5)
    if key not in esperando_pie:
        return
    pending = esperando_pie.pop(key)
    chat_id, msg_id = key
    datos_chat = get_store(chat_id)
    await encolar({
        "image_bytes": pending["image_bytes"], "mime": pending["mime"],
        "pie": pending.get("caption", ""), "chat_id": chat_id,
        "nombre_g": pending["nombre_g"], "origen": "grupo", "msg_id": msg_id,
        "file_id": pending.get("file_id")
    })

# ── Tareas programadas ────────────────────────────────────────────────────────
async def tarea_excel_backup(app):
    """19:50 Argentina — manda Excel del día al admin como respaldo."""
    ahora = now_arg()
    fecha = ahora.strftime("%Y%m%d_%H%M")
    enviados = 0
    for cid, datos in store.items():
        if datos.get("chat_id") == ADMIN_ID:
            continue
        registros = datos.get("registros", [])
        if not registros:
            continue
        nombre = datos.get("nombre", cid)
        try:
            buf = generar_excel(registros, datos["semana_actual"], nombre)
            await app.bot.send_document(
                chat_id=ADMIN_ID,
                document=buf,
                filename=f"Backup_{nombre.replace(' ','_')}_{fecha}.xlsx",
                caption=f"💾 *Respaldo diario — {nombre}*\n📄 {len(registros)} comprobantes | 💰 ${sum(float(r.get('monto') or 0) for r in registros):,.0f} ARS",
                parse_mode="Markdown"
            )
            enviados += 1
        except Exception as e:
            log.error(f"Error backup Excel {cid}: {e}")
    if enviados:
        log.info(f"Backup Excel enviado: {enviados} grupos")

async def tarea_resumen_diario(app):
    fecha_hoy = now_arg().strftime("%d/%m/%Y")
    texto = f"📊 *Resumen diario — {fecha_hoy}*\n\n"
    total_global = 0
    total_comp = 0

    for cid, datos in store.items():
        if datos.get("chat_id") == ADMIN_ID:
            continue
        hoy = [r for r in datos.get("registros", []) if r.get("_fecha_carga","").startswith(fecha_hoy)]
        err = [r for r in datos.get("errores",   []) if r.get("_fecha_carga","").startswith(fecha_hoy)]
        if not hoy and not err:
            continue
        nombre = datos.get("nombre", cid)
        total  = sum(float(r.get("monto") or 0) for r in hoy)
        sin_cvu = sum(1 for r in hoy if not (r.get("cvu_ultimos4") or "").strip())
        total_global += total
        total_comp   += len(hoy)

        texto += f"📍 *{nombre}*\n"
        texto += f"   ✅ {len(hoy)} ok | 💰 ${total:,.0f}"
        if sin_cvu: texto += f" | 🔴 {sin_cvu} sin CVU"
        if err:     texto += f" | ⛔ {len(err)} errores"
        texto += "\n"
        for r in hoy:
            cvu = (r.get("cvu_ultimos4") or "").strip()
            monto_fmt = f"${float(r.get('monto') or 0):,.0f}"
            origen = "📩" if r.get("_origen") == "privado" else "👥"
            texto += f"   {origen} {r.get('remitente','—')} | {monto_fmt} | {'****'+cvu if cvu else '⚠️'}\n"
        texto += "\n"

    if not total_comp:
        texto += "📭 Sin comprobantes hoy."
    else:
        texto += f"─────────────────────\n💰 *TOTAL: ${total_global:,.0f} ARS* ({total_comp} comprobantes)"

    try:
        await app.bot.send_message(chat_id=ADMIN_ID, text=texto, parse_mode="Markdown")
    except Exception as e:
        log.error(f"Error resumen diario: {e}")

async def tarea_excel_semanal(app):
    ahora = now_arg()
    for cid, datos in list(store.items()):
        if datos.get("chat_id") == ADMIN_ID:
            continue
        registros = datos.get("registros", [])
        errores   = datos.get("errores",   [])
        semana    = datos.get("semana_actual", semana_label())
        nombre    = datos.get("nombre", cid)
        fecha     = ahora.strftime("%Y%m%d")

        # Excel normal
        if registros:
            try:
                buf = generar_excel(registros, semana, nombre)
                nombre_arch = f"Comprobantes_{nombre.replace(' ','_')}_{fecha}.xlsx"
                total   = sum(float(r.get("monto") or 0) for r in registros)
                sin_cvu = sum(1 for r in registros if not (r.get("cvu_ultimos4") or "").strip())
                caption = (
                    f"📊 *Excel Semanal — {nombre}*\n"
                    f"📅 {semana} | 📄 {len(registros)} comprobantes | 💰 ${total:,.0f} ARS\n"
                    f"{'⚠️ '+str(sin_cvu)+' sin CVU' if sin_cvu else '✅ Todos con CVU'}"
                )
                await app.bot.send_document(chat_id=int(cid), document=buf, filename=nombre_arch, caption=caption, parse_mode="Markdown")
                buf.seek(0)
                await app.bot.send_document(chat_id=ADMIN_ID, document=buf, filename=nombre_arch, caption=f"📎 {caption}", parse_mode="Markdown")
            except Exception as e:
                log.error(f"Error Excel normal {cid}: {e}")

        # Excel errores + duplicados
        if errores or datos.get("duplicados"):
            try:
                dups = datos.get("duplicados", [])
                buf_err = generar_excel(errores, semana, nombre, es_errores=True, duplicados=dups)
                nombre_err  = f"Rechazados_{nombre.replace(' ','_')}_{fecha}.xlsx"
                caption_err = f"⛔ *Rechazados — {nombre}*\n📅 {semana} | {len(errores)} rechazados | 🔁 {len(dups)} duplicados"
                await app.bot.send_document(chat_id=int(cid), document=buf_err, filename=nombre_err, caption=caption_err, parse_mode="Markdown")
                buf_err.seek(0)
                await app.bot.send_document(chat_id=ADMIN_ID, document=buf_err, filename=nombre_err, caption=f"📎 {caption_err}", parse_mode="Markdown")
            except Exception as e:
                log.error(f"Error Excel errores {cid}: {e}")

        store[cid] = {"nombre": nombre, "semana_actual": semana_label(), "registros": [], "errores": [], "chat_id": int(cid),
                      "ultimo_resumen_idx": 0, "total_mensual": datos.get("total_mensual", 0.0), "mes_actual": datos.get("mes_actual", "")}
    guardar_store()

    try:
        await app.bot.send_message(chat_id=ADMIN_ID,
            text=f"🔄 *Semana reiniciada* — {now_arg().strftime('%d/%m/%Y %H:%M')} hs",
            parse_mode="Markdown")
    except Exception as e:
        log.error(f"Error nueva semana: {e}")

# ── Handlers ──────────────────────────────────────────────────────────────────
async def cmd_start(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "👋 *Bot de Comprobantes Agilpagos*\n\n"
        "Mandame imágenes de comprobantes con los datos al pie.\n\n"
        "📌 *Comandos:*\n"
        "/resumen — comprobantes de este grupo\n"
        "/errores — ver comprobantes rechazados\n"
        "/hoy — resumen de hoy\n"
        "/excel — generar Excel ahora\n"
        "/grupos — todos los grupos (solo admin)\n"
        "/nueva\\_semana — reiniciar semana\n"
        "/borrar — borrar registros\n\n"
        "⏰ *Automático:*\n"
        "📊 Resumen diario → 20:00 hs\n"
        "📎 Excel semanal → Jueves 21:00 hs",
        parse_mode="Markdown"
    )

async def cmd_errores(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    datos = get_store(chat_id, get_nombre_grupo(update))
    errores = datos.get("errores", [])
    if not errores:
        await update.message.reply_text("✅ No hay comprobantes rechazados.")
        return
    texto = f"⛔ *Comprobantes rechazados — {datos['nombre']}*\n─────────────────────\n"
    for i, r in enumerate(errores, 1):
        monto_fmt = f"${float(r.get('monto') or 0):,.0f}"
        texto += f"*#{i}* {r.get('remitente','—')} | {monto_fmt}\n_{r.get('_motivo_error','?')}_\n\n"
    await update.message.reply_text(texto, parse_mode="Markdown")

async def cmd_grupos(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id != ADMIN_ID:
        await update.message.reply_text("⛔ Solo el administrador.")
        return
    if not store:
        await update.message.reply_text("📭 No hay grupos activos.")
        return
    texto = "📋 *Grupos activos:*\n─────────────────────\n"
    total_global = 0
    for cid, datos in store.items():
        if datos.get("chat_id") == ADMIN_ID:
            continue
        regs = datos.get("registros", [])
        errs = datos.get("errores", [])
        total = sum(float(r.get("monto") or 0) for r in regs)
        total_global += total
        texto += f"📍 *{datos.get('nombre',cid)}*\n   ✅ {len(regs)} ok | ⛔ {len(errs)} errores | 💰 ${total:,.0f} ARS\n\n"
    texto += f"💰 *Total global: ${total_global:,.0f} ARS*"
    await update.message.reply_text(texto, parse_mode="Markdown")

async def cmd_hoy(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    datos = get_store(chat_id, get_nombre_grupo(update))
    fecha_hoy = now_arg().strftime("%d/%m/%Y")
    hoy = [r for r in datos.get("registros",[]) if r.get("_fecha_carga","").startswith(fecha_hoy)]
    if not hoy:
        await update.message.reply_text(f"📭 Sin comprobantes hoy ({fecha_hoy}).")
        return
    total = sum(float(r.get("monto") or 0) for r in hoy)
    sin_cvu = sum(1 for r in hoy if not (r.get("cvu_ultimos4") or "").strip())
    detalle = ""
    for i, r in enumerate(hoy, 1):
        cvu = (r.get("cvu_ultimos4") or "").strip()
        detalle += f"\n*#{i}* {r.get('remitente','—')} | ${float(r.get('monto') or 0):,.0f} | {'****'+cvu if cvu else '⚠️'}"
    await update.message.reply_text(
        f"📅 *Hoy {fecha_hoy}*\n📄 {len(hoy)} | 💰 ${total:,.0f} ARS{' | 🔴 '+str(sin_cvu) if sin_cvu else ''}\n─────────────────────{detalle}",
        parse_mode="Markdown"
    )

async def cmd_resumen(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    chat_id    = update.effective_chat.id
    es_privado = update.effective_chat.type == "private"

    if es_privado:
        # Mostrar resumen de TODOS los grupos
        if not store:
            await update.message.reply_text("📭 No hay grupos activos todavía.")
            return
        texto = "📊 *Resumen de todos los grupos*\n─────────────────────\n"
        total_global = 0
        total_comp   = 0
        for cid, datos in store.items():
            if datos.get("chat_id") == ADMIN_ID:
                continue
            regs = datos.get("registros", [])
            errs = datos.get("errores",   [])
            if not regs and not errs:
                texto += f"📍 *{datos.get('nombre', cid)}*: sin comprobantes\n\n"
                continue
            total   = sum(float(r.get("monto") or 0) for r in regs)
            sin_cvu = sum(1 for r in regs if not (r.get("cvu_ultimos4") or "").strip())
            total_global += total
            total_comp   += len(regs)
            texto += (
                f"📍 *{datos.get('nombre', cid)}*\n"
                f"   ✅ Aceptados: {len(regs)}\n"
                f"   ⛔ Rechazados: {len(errs)}\n"
                f"   🔴 Sin CVU: {sin_cvu}\n"
                f"   💰 Total: ${total:,.0f} ARS\n"
                f"   📅 {datos.get('semana_actual','')}\n\n"
            )
        if total_comp > 0:
            texto += f"─────────────────────\n💰 *TOTAL GLOBAL: ${total_global:,.0f} ARS* ({total_comp} comprobantes)"
        await update.message.reply_text(texto, parse_mode="Markdown")
        return

    # Resumen del grupo actual
    datos = get_store(chat_id, get_nombre_grupo(update))
    regs = datos["registros"]
    errs = datos.get("errores", [])
    if not regs and not errs:
        await update.message.reply_text("📭 No hay comprobantes.")
        return
    total   = sum(float(r.get("monto") or 0) for r in regs)
    sin_cvu = sum(1 for r in regs if not (r.get("cvu_ultimos4") or "").strip())
    await update.message.reply_text(
        f"📊 *{datos['nombre']} — {datos['semana_actual']}*\n─────────────────────\n"
        f"✅ Aceptados: {len(regs)}\n⛔ Rechazados: {len(errs)}\n"
        f"🔴 Sin CVU: {sin_cvu}\n💰 Total ARS: ${total:,.0f}",
        parse_mode="Markdown"
    )

async def cmd_excel(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    chat_id    = update.effective_chat.id
    es_privado = update.effective_chat.type == "private"
    args       = ctx.args or []

    # Parsear fechas si se pasan como argumentos: /excel 01/06/2026 12/06/2026
    fecha_desde = None
    fecha_hasta = None
    label_fechas = ""
    if len(args) >= 2:
        try:
            fecha_desde = datetime.strptime(args[0], "%d/%m/%Y").date()
            fecha_hasta = datetime.strptime(args[1], "%d/%m/%Y").date()
            label_fechas = f" ({args[0]} → {args[1]})"
        except ValueError:
            await update.message.reply_text(
                "⚠️ Formato incorrecto. Usá: `/excel DD/MM/YYYY DD/MM/YYYY`\nEjemplo: `/excel 01/06/2026 12/06/2026`",
                parse_mode="Markdown"
            )
            return
    elif len(args) == 1:
        await update.message.reply_text(
            "⚠️ Falta la fecha hasta. Usá: `/excel DD/MM/YYYY DD/MM/YYYY`",
            parse_mode="Markdown"
        )
        return

    def filtrar_por_fecha(registros: list) -> list:
        if not fecha_desde or not fecha_hasta:
            return registros
        resultado = []
        for r in registros:
            incluir = False
            # Intentar por fecha del ticket
            try:
                f = datetime.strptime(r.get("fecha", ""), "%d/%m/%Y").date()
                if fecha_desde <= f <= fecha_hasta:
                    incluir = True
            except Exception:
                pass
            # Si no matchea, intentar por fecha de carga (cuando el bot lo procesó)
            if not incluir:
                try:
                    fc = datetime.strptime(r.get("_fecha_carga", "")[:10], "%d/%m/%Y").date()
                    if fecha_desde <= fc <= fecha_hasta:
                        incluir = True
                except Exception:
                    pass
            if incluir:
                resultado.append(r)
        return resultado

    fecha_archivo = now_arg().strftime("%Y%m%d_%H%M")

    if es_privado:
        enviados = 0
        for cid, datos in store.items():
            if datos.get("chat_id") == ADMIN_ID:
                continue
            nombre    = datos.get("nombre", cid)
            registros = filtrar_por_fecha(datos.get("registros", []))
            errores   = filtrar_por_fecha(datos.get("errores",   []))
            duplicados = filtrar_por_fecha(datos.get("duplicados", []))

            if registros:
                buf = generar_excel(registros, datos["semana_actual"], nombre)
                total = sum(float(r.get("monto") or 0) for r in registros)
                await update.message.reply_document(
                    document=buf,
                    filename=f"Aprobados_{nombre.replace(' ','_')}_{fecha_archivo}.xlsx",
                    caption=f"✅ *{nombre}*{label_fechas}\n{len(registros)} aprobados | ${total:,.0f} ARS",
                    parse_mode="Markdown"
                )
                enviados += 1

            if errores or duplicados:
                buf_err = generar_excel(errores, datos["semana_actual"], nombre, es_errores=True, duplicados=duplicados)
                await update.message.reply_document(
                    document=buf_err,
                    filename=f"Rechazados_{nombre.replace(' ','_')}_{fecha_archivo}.xlsx",
                    caption=f"⛔ *{nombre}*{label_fechas}\n{len(errores)} rechazados | 🔁 {len(duplicados)} duplicados",
                    parse_mode="Markdown"
                )
                enviados += 1

        if not enviados:
            await update.message.reply_text(f"📭 No hay comprobantes{label_fechas}.")
        return

    # Desde el grupo
    datos  = get_store(chat_id, get_nombre_grupo(update))
    nombre = datos.get("nombre", "Grupo")
    registros  = filtrar_por_fecha(datos.get("registros",  []))
    errores    = filtrar_por_fecha(datos.get("errores",    []))
    duplicados = filtrar_por_fecha(datos.get("duplicados", []))
    log.info(f"Excel grupo: {len(registros)} reg, {len(errores)} err, {len(duplicados)} dups (total dups en store: {len(datos.get('duplicados',[]))})")
    enviados = 0

    if registros:
        buf = generar_excel(registros, datos["semana_actual"], nombre)
        total = sum(float(r.get("monto") or 0) for r in registros)
        await update.message.reply_document(
            document=buf,
            filename=f"Aprobados_{nombre.replace(' ','_')}_{fecha_archivo}.xlsx",
            caption=f"✅ *{nombre}*{label_fechas}\n{len(registros)} aprobados | ${total:,.0f} ARS",
            parse_mode="Markdown"
        )
        enviados += 1

    if errores or duplicados or registros:
        buf_err = generar_excel(errores, datos["semana_actual"], nombre, es_errores=True, duplicados=duplicados)
        await update.message.reply_document(
            document=buf_err,
            filename=f"Rechazados_{nombre.replace(' ','_')}_{fecha_archivo}.xlsx",
            caption=f"⛔ *{nombre}*{label_fechas}\n{len(errores)} rechazados | 🔁 {len(duplicados)} duplicados",
            parse_mode="Markdown"
        )
        enviados += 1

    if not enviados:
        await update.message.reply_text(f"📭 No hay comprobantes{label_fechas}.")

async def cmd_nueva_semana(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    kb = [[InlineKeyboardButton("✅ Sí", callback_data="nueva_semana_si"),
           InlineKeyboardButton("❌ No",  callback_data="cancelar")]]
    await update.message.reply_text("⚠️ ¿Iniciar nueva semana? Se borran todos los registros.",
        reply_markup=InlineKeyboardMarkup(kb))

async def cmd_pendientes(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id != ADMIN_ID:
        await update.message.reply_text("Solo el administrador.")
        return
    chat_id = update.effective_chat.id
    es_privado = update.effective_chat.type == "private"
    grupos_a_revisar = list(received_log.keys()) if es_privado else [str(chat_id)]
    total_recibidas = 0
    total_pendientes = 0
    texto = "📋 *Auditoría de fotos*\n─────────────────────\n"
    for cid in grupos_a_revisar:
        entradas = received_log.get(cid, [])
        nombre = store.get(cid, {}).get("nombre", cid)
        recibidas = len(entradas)
        procesadas = [e for e in entradas if e.get("estado") in ("procesada", "procesado")]
        pendientes = [e for e in entradas if e.get("estado") not in ("procesada", "procesado")]
        total_recibidas += recibidas
        total_pendientes += len(pendientes)
        if pendientes:
            texto += f"📍 *{nombre}*\n"
            texto += f"   📥 Recibidas: {recibidas} | ✅ Procesadas: {len(procesadas)} | ⚠️ Sin procesar: {len(pendientes)}\n"
            for e in pendientes[:10]:
                texto += f"   • msg #{e.get('msg_id')} — {e.get('fecha','')} ({e.get('estado','')})\n"
            if len(pendientes) > 10:
                texto += f"   ... y {len(pendientes)-10} más\n"
            texto += "\n"
    if total_pendientes == 0:
        texto += "✅ Todas las fotos fueron procesadas."
    else:
        texto += f"─────────────────────\n⚠️ *Total sin procesar: {total_pendientes} de {total_recibidas}*"
    await update.message.reply_text(texto, parse_mode="Markdown")

async def cmd_limpiar_log(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Limpia el log de auditoría del grupo actual."""
    chat_id = update.effective_chat.id
    cid_str = str(chat_id)
    if cid_str in received_log:
        cantidad = len(received_log[cid_str])
        received_log[cid_str] = []
        guardar_store()
        await update.message.reply_text(f"✅ Log limpiado — {cantidad} entradas eliminadas.")
    else:
        await update.message.reply_text("📭 No había entradas en el log.")

async def cmd_recuperar(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Lee el store.json del Volume y recarga los datos en memoria."""
    if update.effective_user.id != ADMIN_ID:
        await update.message.reply_text("⛔ Solo el administrador.")
        return

    global store
    try:
        if not os.path.exists(DATA_FILE):
            await update.message.reply_text("❌ No existe el archivo /data/store.json en el Volume.")
            return

        with open(DATA_FILE, "r", encoding="utf-8") as f:
            store_recuperado = json.load(f)

        if not store_recuperado:
            await update.message.reply_text("❌ El archivo existe pero está vacío.")
            return

        store = store_recuperado
        total_registros = sum(len(d.get("registros", [])) for d in store.values())
        total_errores   = sum(len(d.get("errores",   [])) for d in store.values())
        total_grupos    = len(store)

        texto = f"✅ *Store recuperado desde Volume*\n─────────────────────\n"
        texto += f"📍 Grupos: {total_grupos}\n"
        texto += f"✅ Registros: {total_registros}\n"
        texto += f"⛔ Errores: {total_errores}\n\n"

        for cid, datos in store.items():
            regs = datos.get("registros", [])
            errs = datos.get("errores", [])
            total = sum(float(r.get("monto") or 0) for r in regs)
            texto += f"📍 *{datos.get('nombre', cid)}*: {len(regs)} ok | {len(errs)} err | ${total:,.0f} ARS\n"

        await update.message.reply_text(texto, parse_mode="Markdown")

    except Exception as e:
        await update.message.reply_text(f"❌ Error al recuperar: {e}")

async def cmd_borrar(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    kb = [[InlineKeyboardButton("✅ Sí", callback_data="borrar_si"),
           InlineKeyboardButton("❌ No",  callback_data="cancelar")]]
    await update.message.reply_text("⚠️ ¿Borrar TODOS los comprobantes (incluso errores)?",
        reply_markup=InlineKeyboardMarkup(kb))

async def callback_handler(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    chat_id = update.effective_chat.id
    user_id = update.effective_user.id
    data    = query.data

    if data.startswith("asignar_grupo_"):
        grupo_cid = data.replace("asignar_grupo_", "")
        if user_id not in pendientes:
            await query.edit_message_text("⏰ La foto expiró. Mandala de nuevo.")
            return
        pending = pendientes.pop(user_id)
        await query.edit_message_text("🔍 Analizando comprobante con IA...")
        try:
            resultado = await analizar_imagen(pending["image_bytes"], pending["mime"])
            datos_g   = get_store(int(grupo_cid))
            nombre_g  = datos_g["nombre"]
            coincide, motivo = verificar_pie(resultado, pending.get("caption",""))
            num = len(datos_g["registros"]) + len(datos_g.get("errores",[])) + 1
            resultado.update({"_num": num, "_fecha_carga": now_arg().strftime("%d/%m/%Y %H:%M"),
                               "_origen": "privado", "_pie": pending.get("caption","")})
            if coincide:
                datos_g["registros"].append(resultado)
                texto = "📩 *Recibido por privado*\n" + formatear_resultado(resultado, num, nombre_g)
                cvu = (resultado.get("cvu_ultimos4") or "").strip()
                kb = None
                if not cvu:
                    kb = InlineKeyboardMarkup([[InlineKeyboardButton(f"✏️ CVU #{num}", callback_data=f"editar_cvu_{num}")]])
                await query.edit_message_text(texto, parse_mode="Markdown", reply_markup=kb)
            else:
                resultado["_motivo_error"] = motivo
                datos_g.setdefault("errores", []).append(resultado)
                await query.edit_message_text(
                    f"⛔ *Comprobante #{num} RECHAZADO*\n❌ {motivo}\n_Guardado en errores._",
                    parse_mode="Markdown")
        except Exception as e:
            await query.edit_message_text(f"❌ Error: {e}")
        return

    if data == "cancelar_privado":
        pendientes.pop(user_id, None)
        await query.edit_message_text("❌ Comprobante descartado.")
        return

    if data == "nueva_semana_si":
        n  = store.get(str(chat_id), {}).get("nombre", "Grupo")
        tm = store.get(str(chat_id), {}).get("total_mensual", 0.0)
        ma = store.get(str(chat_id), {}).get("mes_actual", "")
        datos_actuales = store.get(str(chat_id), {})
        registros = datos_actuales.get("registros", [])
        errores   = datos_actuales.get("errores", [])
        duplicados = datos_actuales.get("duplicados", [])
        semana    = datos_actuales.get("semana_actual", semana_label())
        fecha_arch = now_arg().strftime("%Y%m%d_%H%M")

        # ── Backup automático antes de borrar ──
        if registros:
            try:
                buf = generar_excel(registros, semana, n)
                total = sum(float(r.get("monto") or 0) for r in registros)
                sin_cvu = sum(1 for r in registros if not (r.get("cvu_ultimos4") or "").strip())
                await query.message.reply_document(
                    document=buf,
                    filename=f"BACKUP_antes_nueva_semana_{n.replace(' ','_')}_{fecha_arch}.xlsx",
                    caption=(
                        f"💾 *Backup automático — ANTES de nueva semana*\n"
                        f"📍 {n} | 📅 {semana}\n"
                        f"📄 {len(registros)} aprobados | 💰 ${total:,.0f} ARS\n"
                        f"{'⚠️ '+str(sin_cvu)+' sin CVU' if sin_cvu else '✅ Todos con CVU'}"
                    ),
                    parse_mode="Markdown"
                )
            except Exception as e:
                log.error(f"Error backup nueva semana: {e}")

        if errores or duplicados:
            try:
                buf_err = generar_excel(errores, semana, n, es_errores=True, duplicados=duplicados)
                await query.message.reply_document(
                    document=buf_err,
                    filename=f"BACKUP_rechazados_{n.replace(' ','_')}_{fecha_arch}.xlsx",
                    caption=f"⛔ *Backup rechazados — {n}*\n{len(errores)} rechazados | 🔁 {len(duplicados)} duplicados",
                    parse_mode="Markdown"
                )
            except Exception as e:
                log.error(f"Error backup errores nueva semana: {e}")

        # También mandar al admin
        try:
            if registros:
                buf2 = generar_excel(registros, semana, n)
                await query.get_bot().send_document(
                    chat_id=ADMIN_ID,
                    document=buf2,
                    filename=f"BACKUP_antes_nueva_semana_{n.replace(' ','_')}_{fecha_arch}.xlsx",
                    caption=f"💾 *Backup admin — nueva semana {n}*\n{len(registros)} registros | ${sum(float(r.get('monto') or 0) for r in registros):,.0f} ARS",
                    parse_mode="Markdown"
                )
        except Exception as e:
            log.error(f"Error enviando backup al admin: {e}")

        store[str(chat_id)] = {"nombre": n, "semana_actual": semana_label(), "registros": [],
                                "errores": [], "chat_id": chat_id,
                                "ultimo_resumen_idx": 0, "total_mensual": tm, "mes_actual": ma}
        guardar_store()
        await query.edit_message_text("✅ Nueva semana iniciada. Backup enviado arriba 👆")
    elif data == "borrar_si":
        d = get_store(chat_id)
        d["registros"] = []
        d["errores"]   = []
        await query.edit_message_text("🗑 Todo borrado.")
    elif data == "cancelar":
        await query.edit_message_text("❌ Cancelado.")

def es_duplicado(resultado: dict, registros: list) -> bool:
    """Detecta duplicados SOLO por número de operación/ID único.
    Si no hay ID, acepta el comprobante directamente.
    """
    ref_nuevo = (resultado.get("referencia") or "").strip()

    # Sin ID de operación → no se puede verificar duplicado → aceptar
    if not ref_nuevo:
        return False

    # Comparar contra todos los registros existentes
    for r in registros:
        ref = (r.get("referencia") or "").strip()
        if ref and ref == ref_nuevo:
            return True
    return False

async def obtener_imagen(update, ctx) -> tuple:
    """Extrae imagen de cualquier tipo de mensaje (foto, documento, reenvío)."""
    msg = update.message
    if msg.photo:
        file = await ctx.bot.get_file(msg.photo[-1].file_id)
        async with httpx.AsyncClient() as client:
            return (await client.get(file.file_path)).content, "image/jpeg"
    elif msg.document and msg.document.mime_type and msg.document.mime_type.startswith("image/"):
        file = await ctx.bot.get_file(msg.document.file_id)
        async with httpx.AsyncClient() as client:
            return (await client.get(file.file_path)).content, msg.document.mime_type
    return None, None

# ── Worker de cola de procesamiento ──────────────────────────────────────────
REDIS_QUEUE_KEY = "bot_cola_comprobantes"

async def encolar(item: dict):
    """Agrega un item a Redis o a la cola local."""
    global redis_client
    try:
        if redis_client:
            # Guardar sin image_bytes (se re-descarga desde Telegram)
            item_redis = {k: v for k, v in item.items() if k != "image_bytes"}
            await redis_client.rpush(REDIS_QUEUE_KEY, json.dumps(item_redis, ensure_ascii=False))
            return
    except Exception as e:
        log.error(f"Error encolando en Redis: {e}")
    # Fallback: cola local
    await cola_procesamiento.put(item)

async def worker_procesamiento(app):
    """Procesa comprobantes de Redis o cola local."""
    global redis_client
    while True:
        try:
            item = None
            # Intentar desde Redis primero
            if redis_client:
                try:
                    result = await redis_client.blpop(REDIS_QUEUE_KEY, timeout=1)
                    if result:
                        item_data = json.loads(result[1])
                        # Re-descargar imagen desde Telegram
                        file_id = item_data.get("file_id")
                        if file_id:
                            descargado = False
                            for intento in range(3):
                                try:
                                    file = await app.bot.get_file(file_id)
                                    async with httpx.AsyncClient() as client:
                                        resp = await client.get(file.file_path)
                                        item_data["image_bytes"] = resp.content
                                        item_data["mime"] = item_data.get("mime", "image/jpeg")
                                        item = item_data
                                        descargado = True
                                        break
                                except Exception as e:
                                    log.warning(f"Descarga intento {intento+1}/3 fallida {file_id}: {e}")
                                    await asyncio.sleep(2)
                            if not descargado:
                                reintentos = item_data.get("_reintentos", 0) + 1
                                if reintentos <= 3:
                                    item_data["_reintentos"] = reintentos
                                    await redis_client.rpush(REDIS_QUEUE_KEY, json.dumps(item_data))
                                    log.warning(f"Reencolado (intento {reintentos}/3): msg {item_data.get('msg_id')}")
                                else:
                                    log.error(f"Descartado tras 3 reintentos: msg {item_data.get('msg_id')}")
                                    try:
                                        await app.bot.send_message(chat_id=item_data["chat_id"], text="⚠️ No se pudo procesar este comprobante tras 3 intentos. Reenvialo.", reply_to_message_id=item_data.get("msg_id"))
                                        asyncio.create_task(reaccionar(app.bot, item_data["chat_id"], item_data["msg_id"], "🤡"))
                                    except Exception as e:
                                        log.error(f"Error notificando descarte: {e}")
                                continue
                        else:
                            continue
                except Exception as e:
                    log.error(f"Error leyendo Redis: {e}")
                    await asyncio.sleep(1)
                    continue
            else:
                # Fallback: cola local
                try:
                    item = await asyncio.wait_for(cola_procesamiento.get(), timeout=1)
                except asyncio.TimeoutError:
                    continue

            if item and item.get("image_bytes"):
                try:
                    await procesar_comprobante(
                        item["image_bytes"], item["mime"], item.get("pie", ""),
                        item["chat_id"], item["nombre_g"], item.get("origen", "grupo"),
                        app.bot, item["msg_id"]
                    )
                except Exception as e:
                    log.error(f"Error en worker: {e}")
                finally:
                    if not redis_client:
                        cola_procesamiento.task_done()
        except Exception as e:
            log.error(f"Error en worker loop: {e}")
            await asyncio.sleep(1)

# ── Soporte para álbumes (media groups) ──────────────────────────────────────
media_groups: dict = {}  # {(chat_id, media_group_id): {"images": [], "caption": "", "nombre_g": "", "timer": task}}

async def procesar_album(key, app):
    """Espera 3 segundos para que lleguen todas las fotos del álbum, luego las procesa."""
    await asyncio.sleep(3)
    if key not in media_groups:
        return
    grupo = media_groups.pop(key)
    chat_id, media_group_id = key
    nombre_g = grupo["nombre_g"]
    caption  = grupo.get("caption", "")
    images   = grupo["images"]

    for image_bytes, mime, msg_id, file_id in images:
        await encolar({"image_bytes": image_bytes, "mime": mime, "pie": caption, "chat_id": chat_id, "nombre_g": nombre_g, "origen": "grupo", "msg_id": msg_id, "file_id": file_id})

async def handle_photo(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    chat_id    = update.effective_chat.id
    es_privado = update.effective_chat.type == "private"
    caption    = (update.message.caption or "").strip()
    msg        = update.message
    msg_id     = msg.message_id

    log.info(f"📥 Foto recibida — chat:{chat_id} msg:{msg_id} album:{msg.media_group_id or 'no'}")
    if not es_privado:
        cid_str = str(chat_id)
        if cid_str not in received_log:
            received_log[cid_str] = []
        received_log[cid_str].append({"msg_id": msg_id, "fecha": now_arg().strftime("%d/%m/%Y %H:%M"), "nombre_g": get_nombre_grupo(update), "estado": "recibida", "album": bool(msg.media_group_id)})

    image_bytes, mime = await obtener_imagen(update, ctx)
    if not image_bytes:
        log.warning(f"⚠️ No se pudo obtener imagen — chat:{chat_id} msg:{msg_id}")
        return

    if es_privado:
        grupos = grupos_disponibles()
        if not grupos:
            await update.message.reply_text("⚠️ No hay grupos activos todavía.")
            return
        pendientes[update.effective_user.id] = {"image_bytes": image_bytes, "mime": mime, "caption": caption}
        botones = [[InlineKeyboardButton(f"📍 {n}", callback_data=f"asignar_grupo_{cid}")] for cid, n in grupos]
        botones.append([InlineKeyboardButton("❌ Cancelar", callback_data="cancelar_privado")])
        await update.message.reply_text("*¿A qué grupo pertenece este comprobante?*",
            reply_markup=InlineKeyboardMarkup(botones), parse_mode="Markdown")
        return

    nombre_g = get_nombre_grupo(update)

    # ── Álbum (media group) ──
    if msg.media_group_id:
        key = (chat_id, msg.media_group_id)
        if key not in media_groups:
            media_groups[key] = {"images": [], "caption": caption, "nombre_g": nombre_g}
        else:
            # Si el nuevo mensaje tiene caption, usarlo
            if caption:
                media_groups[key]["caption"] = caption

        file_id = msg.photo[-1].file_id if msg.photo else (msg.document.file_id if msg.document else None)
        media_groups[key]["images"].append((image_bytes, mime, msg_id, file_id))

        # Cancelar timer anterior y crear uno nuevo
        if "timer" in media_groups[key]:
            media_groups[key]["timer"].cancel()
        task = asyncio.create_task(procesar_album(key, ctx.application))
        media_groups[key]["timer"] = task
        return

    # ── Imagen individual ──
    # Procesar directo — con o sin pie (ya no es obligatorio)
    if caption:
        # Guardar file_id para poder re-descargar si el bot reinicia
        file_id = msg.photo[-1].file_id if msg.photo else (msg.document.file_id if msg.document else None)
        await encolar({
            "image_bytes": image_bytes, "mime": mime, "pie": caption,
            "chat_id": chat_id, "nombre_g": nombre_g, "origen": "grupo", "msg_id": msg_id,
            "file_id": file_id
        })
    else:
        key = (chat_id, msg_id)
        file_id_sin_pie = msg.photo[-1].file_id if msg.photo else (msg.document.file_id if msg.document else None)
        esperando_pie[key] = {"image_bytes": image_bytes, "mime": mime, "caption": "", "nombre_g": nombre_g, "file_id": file_id_sin_pie}
        task = asyncio.create_task(procesar_sin_pie(key, ctx.application))
        esperando_pie[key]["task"] = task

async def handle_document(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    doc = update.message.document
    if not doc or not doc.mime_type or not doc.mime_type.startswith("image/"):
        return
    # Reusar handle_photo ya que tiene la misma lógica
    await handle_photo(update, ctx)

async def handle_text(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Captura texto del pie, correcciones a rechazos, resúmenes de tanda y totales mensuales."""
    if update.effective_chat.type == "private":
        return
    chat_id = update.effective_chat.id
    texto   = (update.message.text or "").strip()
    if not texto:
        return

    texto_lower = texto.lower()

    # ── Detectar resumen de tanda: contiene "tickets" + número + monto ──
    import re
    es_tanda = "tickets" in texto_lower
    es_total_mes = re.search(r"total\s+\w+\s*:", texto_lower)

    if es_tanda and not update.message.reply_to_message:
        datos = get_store(chat_id, get_nombre_grupo(update))
        registros = datos.get("registros", [])
        idx_desde = datos.get("ultimo_resumen_idx", 0)

        # Comprobantes de esta tanda (desde el último resumen)
        tanda = registros[idx_desde:]

        # Extraer número de tickets del mensaje
        num_match    = re.search(r"tickets[:\s]*(\d[\d.,]*)", texto_lower)
        monto_match  = re.search(r"\$\s*([\d.,]+)", texto)

        tickets_msg = int(num_match.group(1).replace(".", "").replace(",", "")) if num_match else None
        monto_msg   = float(monto_match.group(1).replace(".", "").replace(",", "")) if monto_match else None

        # Calcular real
        tickets_real = len(tanda)
        monto_real   = sum(float(r.get("monto") or 0) for r in tanda)

        # Duplicados en esta tanda
        duplicados_tanda = [r for r in datos.get("duplicados", [])
                           if r.get("_fecha_carga","") >= (tanda[0].get("_fecha_carga","") if tanda else "")]

        # Actualizar índice de último resumen
        datos["ultimo_resumen_idx"] = len(registros)
        # NO limpiar duplicados — se acumulan hasta nueva semana para el Excel

        # Acumular al total mensual
        datos["total_mensual"] = datos.get("total_mensual", 0.0) + monto_real
        guardar_store()

        # Comparar
        tickets_ok = tickets_msg is None or tickets_msg == tickets_real
        monto_ok   = monto_msg is None or abs(monto_msg - monto_real) < 1

        dup_line = f"\n🔁 Duplicados ignorados: *{len(duplicados_tanda)}*" if duplicados_tanda else ""

        if tickets_ok and monto_ok:
            await update.message.reply_text(
                f"✅ *Tanda verificada — todo coincide*\n"
                f"🎫 Tickets: {tickets_real}\n"
                f"💰 Monto: ${monto_real:,.0f} ARS{dup_line}",
                parse_mode="Markdown"
            )
        else:
            diferencias = ""
            if not tickets_ok:
                diferencias += f"🎫 Tickets: informaron *{tickets_msg}* → real *{tickets_real}*\n"
            if not monto_ok:
                diferencias += f"💰 Monto: informaron *${monto_msg:,.0f}* → real *${monto_real:,.0f}*\n"
            await update.message.reply_text(
                f"⚠️ *Tanda con diferencias*\n"
                f"─────────────────────\n"
                f"{diferencias}{dup_line}\n"
                f"_Los datos correctos del bot son los indicados arriba._",
                parse_mode="Markdown"
            )
        return

    # ── Detectar total mensual: "TOTAL JUNIO: $X" ──
    if es_total_mes:
        datos = get_store(chat_id, get_nombre_grupo(update))
        monto_match = re.search(r"\$\s*([\d.,]+)", texto)
        monto_msg   = float(monto_match.group(1).replace(".", "").replace(",", "")) if monto_match else None
        total_real  = datos.get("total_mensual", 0.0)

        # Extraer nombre del mes del mensaje
        mes_match = re.search(r"total\s+(\w+)\s*:", texto_lower)
        nombre_mes = mes_match.group(1).upper() if mes_match else "MES"

        if monto_msg is None:
            await update.message.reply_text(
                f"📅 *Total acumulado {nombre_mes}*\n💰 ${total_real:,.0f} ARS",
                parse_mode="Markdown"
            )
        elif abs(monto_msg - total_real) < 1:
            await update.message.reply_text(
                f"✅ *Total {nombre_mes} verificado*\n💰 ${total_real:,.0f} ARS — coincide ✓",
                parse_mode="Markdown"
            )
        else:
            await update.message.reply_text(
                f"⚠️ *Total {nombre_mes} con diferencia*\n"
                f"Informaron: *${monto_msg:,.0f}*\n"
                f"Real acumulado: *${total_real:,.0f}*\n"
                f"Diferencia: *${abs(monto_msg - total_real):,.0f}*",
                parse_mode="Markdown"
            )
        return

    # ── Detectar si es respuesta a un mensaje de rechazo ──
    reply = update.message.reply_to_message
    if reply:
        key_rechazo = (chat_id, reply.message_id)
        if key_rechazo in mensajes_rechazo:
            info    = mensajes_rechazo[key_rechazo]
            num     = info["num"]
            datos   = get_store(chat_id, get_nombre_grupo(update))
            errores = datos.get("errores", [])

            error_idx = next((i for i, r in enumerate(errores) if r.get("_num") == num), None)
            if error_idx is None:
                await update.message.reply_text("⚠️ No encontré el comprobante a corregir.")
                return

            registro = errores[error_idx]
            coincide, motivo = verificar_pie(registro, texto)

            if coincide:
                errores.pop(error_idx)
                registro["_pie"]          = texto
                registro["_corregido"]    = True
                registro["_motivo_error"] = ""
                if not registro.get("tiene_remitente"):
                    registro["remitente"] = texto
                datos["registros"].append(registro)
                del mensajes_rechazo[key_rechazo]
                guardar_store()

                monto_fmt = f"${float(registro.get('monto') or 0):,.0f}"
                cvu = (registro.get("cvu_ultimos4") or "").strip()
                # Confirmar en el grupo brevemente
                await update.message.reply_text(f"✅ Comprobante #{num} corregido y aceptado.")
                # Detalle completo al admin
                try:
                    await ctx.bot.send_message(
                        chat_id=ADMIN_ID,
                        text=(
                            f"✅ *Comprobante #{num} corregido — {datos.get('nombre','')}*\n"
                            f"💰 {monto_fmt} | 🏦 {'****'+cvu if cvu else '⚠️ Sin CVU'}\n"
                            f"📝 Datos corregidos: _{texto}_"
                        ),
                        parse_mode="Markdown"
                    )
                except Exception:
                    pass
            else:
                await update.message.reply_text(
                    f"⛔ Sigue sin coincidir. Revisá los datos e intentá de nuevo respondiendo a este mensaje.",
                )
            return

    # ── Detectar texto del pie para foto esperando ──
    if texto.count("\n") >= 2:
        return

    key_match = None
    for key in list(esperando_pie.keys()):
        if key[0] == chat_id:
            key_match = key
            break

    if not key_match:
        return

    pending = esperando_pie.pop(key_match)
    task = pending.get("task")
    if task:
        task.cancel()

    await procesar_comprobante(
        pending["image_bytes"], pending["mime"], texto,
        chat_id, pending["nombre_g"], "grupo", ctx.bot, key_match[1]
    )

# ── Main ──────────────────────────────────────────────────────────────────────
async def handle_any(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Handler de respaldo — captura cualquier mensaje con foto no procesado."""
    if not update.message:
        return
    msg = update.message
    # Solo procesar si tiene foto o documento imagen y no fue procesado antes
    tiene_foto = bool(msg.photo)
    tiene_doc  = bool(msg.document and msg.document.mime_type and msg.document.mime_type.startswith("image/"))
    if not tiene_foto and not tiene_doc:
        return
    log.info(f"handle_any capturó mensaje con foto — chat:{update.effective_chat.id} msg:{msg.message_id}")
    await handle_photo(update, ctx)

async def health_server():
    """Servidor HTTP mínimo para health check de Railway."""
    import asyncio
    from http.server import BaseHTTPRequestHandler, HTTPServer
    import threading

    class Handler(BaseHTTPRequestHandler):
        def do_GET(self):
            self.send_response(200)
            self.end_headers()
            self.wfile.write(b"OK")
        def log_message(self, format, *args):
            pass  # Silenciar logs

    port = int(os.environ.get("PORT", 8080))
    server = HTTPServer(("0.0.0.0", port), Handler)
    thread = threading.Thread(target=server.serve_forever, daemon=True)
    thread.start()
    log.info(f"Health server en puerto {port}")

async def main_async():
    cargar_store()

    # Iniciar health server
    try:
        await health_server()
    except Exception as e:
        log.warning(f"Health server no disponible: {e}")

    app = Application.builder().token(TELEGRAM_TOKEN).build()

    app.add_handler(CommandHandler("start",        cmd_start))
    app.add_handler(CommandHandler("hoy",          cmd_hoy))
    app.add_handler(CommandHandler("resumen",      cmd_resumen))
    app.add_handler(CommandHandler("errores",      cmd_errores))
    app.add_handler(CommandHandler("excel",        cmd_excel))
    app.add_handler(CommandHandler("grupos",       cmd_grupos))
    app.add_handler(CommandHandler("nueva_semana", cmd_nueva_semana))
    app.add_handler(CommandHandler("borrar",       cmd_borrar))
    app.add_handler(CommandHandler("recuperar",    cmd_recuperar))
    app.add_handler(CommandHandler("pendientes",   cmd_pendientes))
    app.add_handler(CommandHandler("limpiar_log",  cmd_limpiar_log))
    app.add_handler(CallbackQueryHandler(callback_handler))
    app.add_handler(MessageHandler(filters.PHOTO,          handle_photo))
    app.add_handler(MessageHandler(filters.Document.IMAGE, handle_document))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_text))
    app.add_handler(MessageHandler(filters.ALL, handle_any))

    scheduler = AsyncIOScheduler(timezone="America/Argentina/Buenos_Aires")
    scheduler.add_job(tarea_resumen_diario, CronTrigger(hour=20, minute=0, timezone="America/Argentina/Buenos_Aires"), args=[app], id="resumen_diario")
    scheduler.add_job(tarea_excel_semanal,  CronTrigger(day_of_week="thu", hour=21, minute=0, timezone="America/Argentina/Buenos_Aires"), args=[app], id="excel_semanal")
    scheduler.add_job(tarea_excel_backup, CronTrigger(hour=19, minute=50, timezone="America/Argentina/Buenos_Aires"), args=[app], id="excel_backup")
    scheduler.start()

    # Iniciar Redis
    global redis_client
    if REDIS_URL:
        try:
            redis_client = await aioredis.from_url(REDIS_URL, decode_responses=True)
            await redis_client.ping()
            pending_count = await redis_client.llen(REDIS_QUEUE_KEY)
            log.info(f"Redis conectado ✅ — {pending_count} items pendientes en cola")
        except Exception as e:
            log.error(f"Error conectando Redis: {e}")
            redis_client = None

    log.info("🤖 Bot iniciado con verificación de pie")
    # Iniciar workers de procesamiento (10 workers paralelos)
    for _ in range(10):
        asyncio.create_task(worker_procesamiento(app))
    await app.initialize()
    await app.start()
    await app.updater.start_polling(drop_pending_updates=True)

    # Mantener corriendo
    try:
        await asyncio.Event().wait()
    finally:
        await app.updater.stop()
        await app.stop()
        await app.shutdown()

def main():
    asyncio.run(main_async())

if __name__ == "__main__":
    main()
